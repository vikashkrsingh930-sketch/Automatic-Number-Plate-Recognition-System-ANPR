from ultralytics import YOLO
import cv2
import re
from datetime import datetime
import os
from transformers import TrOCRProcessor, VisionEncoderDecoderModel
from PIL import Image
from openpyxl import Workbook, load_workbook

# Load models
model = YOLO("model/best.pt")
processor = TrOCRProcessor.from_pretrained("microsoft/trocr-base-printed")
ocr_model = VisionEncoderDecoderModel.from_pretrained("microsoft/trocr-base-printed")

# Paths
base_dir = os.path.dirname(__file__)
excel_file = os.path.join(base_dir, "plates.xlsx")
image_dir = os.path.join(base_dir, "plates_images")

if not os.path.exists(image_dir):
    os.makedirs(image_dir)

# Excel setup
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "ANPR Logs"
    ws.append(["Plate Number", "Date", "Time", "Image Path"])
    wb.save(excel_file)
else:
    wb = load_workbook(excel_file)
    ws = wb.active

print("📁 Logging to:", excel_file)

plate_memory = {}
saved_plates = set()


# Clean plate
def clean_plate(text):
    text = text.upper()
    text = re.sub(r'[^A-Z0-9]', '', text)

    text = list(text)
    for i, c in enumerate(text):
        if i < 2:
            if c == '0':
                text[i] = 'O'
        elif 2 <= i < 4:
            if c == 'O':
                text[i] = '0'
        else:
            if c == 'O':
                text[i] = '0'
            if c == 'I':
                text[i] = '1'

    text = "".join(text)
    match = re.findall(r'[A-Z]{2}[0-9]{2}[A-Z]{1,2}[0-9]{4}', text)

    return match[0] if match else ""


# Save data
def save_data(plate, plate_img):
    if plate in saved_plates:
        return

    saved_plates.add(plate)

    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H-%M-%S")

    image_name = f"{plate}_{time}.jpg"
    image_path = os.path.join(image_dir, image_name)

    cv2.imwrite(image_path, plate_img)

    ws.append([plate, date, time, image_path])
    wb.save(excel_file)

    print("📸 IMAGE SAVED:", image_name)
    print("📊 EXCEL SAVED:", plate)


# OCR function
def process_plate(plate):
    try:
        gray = cv2.cvtColor(plate, cv2.COLOR_BGR2GRAY)
        gray = cv2.resize(gray, None, fx=3, fy=3)

        rgb = cv2.cvtColor(gray, cv2.COLOR_GRAY2RGB)
        pil_img = Image.fromarray(rgb)

        pixel_values = processor(images=pil_img, return_tensors="pt").pixel_values
        generated_ids = ocr_model.generate(pixel_values)
        text = processor.batch_decode(generated_ids, skip_special_tokens=True)[0]

        print("OCR:", text)

        cleaned = clean_plate(text)

        if cleaned:
            plate_memory[cleaned] = plate_memory.get(cleaned, 0) + 1

            if plate_memory[cleaned] >= 3:
                save_data(cleaned, plate)
                plate_memory[cleaned] = 0

    except Exception as e:
        print("OCR ERROR:", e)


# Processing function
def process_frame(frame):
    results = model(frame, conf=0.25)

    for box in results[0].boxes:
        x1, y1, x2, y2 = map(int, box.xyxy[0])

        pad = 20
        x1 = max(0, x1 - pad)
        y1 = max(0, y1 - pad)
        x2 = min(frame.shape[1], x2 + pad)
        y2 = min(frame.shape[0], y2 + pad)

        plate = frame[y1:y2, x1:x2]

        if plate.size == 0:
            continue

        process_plate(plate)

        cv2.rectangle(frame, (x1, y1), (x2, y2), (0,255,0), 2)

    return frame


# 🔥 MODE SELECTION
print("\nSelect Mode:")
print("1 - Webcam")
print("2 - Image")
print("3 - Video")

choice = input("Enter choice: ")

# 📷 WEBCAM
if choice == "1":
    cap = cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        frame = process_frame(frame)
        cv2.imshow("Webcam ANPR", frame)

        if cv2.waitKey(1) == 27:
            break

    cap.release()


# 🖼 IMAGE
elif choice == "2":
    path = input("Enter image path: ")
    frame = cv2.imread(path)

    frame = process_frame(frame)
    cv2.imshow("Image ANPR", frame)
    cv2.waitKey(0)


# 🎥 VIDEO
elif choice == "3":
    path = input("Enter video path: ")
    cap = cv2.VideoCapture(path)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        frame = process_frame(frame)
        cv2.imshow("Video ANPR", frame)

        if cv2.waitKey(1) == 27:
            break

    cap.release()

cv2.destroyAllWindows()